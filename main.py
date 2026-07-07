import time
import re
import os
import pyautogui
import pyperclip
import openpyxl
import keyboard

# =========================
# CONFIGURATION
# =========================

EXCEL_FILE = "contacts.xlsx"
SABRE_COMMAND = "*H9*PE*HI\n"

# Pixel coordinates inside the dark blue Sabre area
SABRE_X1, SABRE_Y1 = 10, 100
SABRE_X2, SABRE_Y2 = 1700, 980

SAVE_EVERY = 10
SAVE_RETRIES = 6
SAVE_FAIL_LIMIT = 3

STOP_HOTKEY = "ctrl+alt+k"

# =========================
# STOP CONTROL
# =========================

STOP_REQUESTED = False

def request_stop():
    global STOP_REQUESTED
    STOP_REQUESTED = True
    print(f"Stop requested ({STOP_HOTKEY}). Finishing current step and saving...")

keyboard.add_hotkey(STOP_HOTKEY, request_stop)

# =========================
# SAFE SAVE (atomic replace)
# =========================

def safe_save_workbook(workbook, filename, retries=SAVE_RETRIES):
    temp_name = filename + ".tmp.xlsx"
    for _ in range(retries):
        try:
            workbook.save(temp_name)
            os.replace(temp_name, filename)
            return True
        except Exception:
            time.sleep(0.6)
        finally:
            try:
                if os.path.exists(temp_name):
                    os.remove(temp_name)
            except Exception:
                pass
    return False

# =========================
# EXCEL SETUP
# =========================

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
except FileNotFoundError:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

PHONES_SHEET = "PHONES"
EMAILS_SHEET = "EMAILS"

if PHONES_SHEET not in wb.sheetnames:
    wb.create_sheet(PHONES_SHEET)
if EMAILS_SHEET not in wb.sheetnames:
    wb.create_sheet(EMAILS_SHEET)

ws_phones = wb[PHONES_SHEET]
ws_emails = wb[EMAILS_SHEET]

def next_row(ws):
    return ws.max_row + 1 if ws.max_row > 1 else 1

phones_next_row = next_row(ws_phones)
emails_next_row = next_row(ws_emails)

seen_phones = set()
seen_emails = set()

# Load existing phones
for r in range(1, ws_phones.max_row + 1):
    v = ws_phones.cell(r, 1).value
    if v is not None:
        d = re.sub(r"\D", "", str(v))
        if len(d) == 10:
            seen_phones.add(d)

# Load existing emails
for r in range(1, ws_emails.max_row + 1):
    v = ws_emails.cell(r, 1).value
    if v is not None:
        e = str(v).strip().lower()
        if e:
            seen_emails.add(e)

write_counter = 0
save_failures = 0

def maybe_save():
    global save_failures
    ok = safe_save_workbook(wb, EXCEL_FILE)
    if ok:
        save_failures = 0
        print(f"Saved after {write_counter} new items")
    else:
        save_failures += 1
        print("Save failed. Keep contacts.xlsx closed while running.")
        if save_failures >= SAVE_FAIL_LIMIT:
            raise SystemExit("Stopping due to repeated save failures")

# =========================
# SABRE COPY (drag-select)
# =========================

def copy_sabre_screen():
    pyautogui.click(SABRE_X1 + 10, SABRE_Y1 + 10)
    time.sleep(0.15)

    pyautogui.moveTo(SABRE_X1, SABRE_Y1)
    pyautogui.dragTo(SABRE_X2, SABRE_Y2, duration=0.25, button="left")

    time.sleep(0.1)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(0.25)
    return pyperclip.paste()

# =========================
# PHONE EXTRACTION (A9 lines)
# =========================

A9_LINE_REGEX = re.compile(r"^\s*A9\s+(.*)$", re.MULTILINE)
PHONE_FLEX_REGEX = re.compile(r"(?<!\d)(?:1\D*)?(?:\d\D*){10}(?!\d)")

def normalize_phone_from_match(match_text):
    digits = re.sub(r"\D", "", match_text)
    if len(digits) == 10:
        return digits
    if len(digits) == 11 and digits.startswith("1"):
        return digits[1:]
    return None

def extract_phones(page):
    phones = []
    seen_local = set()

    for m in A9_LINE_REGEX.finditer(page):
        a9_text = m.group(1)
        for pm in PHONE_FLEX_REGEX.finditer(a9_text):
            p = normalize_phone_from_match(pm.group(0))
            if p and p not in seen_local:
                phones.append(p)
                seen_local.add(p)

    return phones

# =========================
# EMAIL EXTRACTION (¥...¥ first, fallback to plain emails)
# =========================

YEN_EMAIL_REGEX = re.compile(r"¥([^¥]+)¥")
PLAIN_EMAIL_REGEX = re.compile(r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b")

def extract_emails(page):
    emails = []
    seen_local = set()

    # Primary: emails wrapped in ¥ ... ¥
    for raw in YEN_EMAIL_REGEX.findall(page):
        e = raw.strip().lower()
        if e and e not in seen_local and PLAIN_EMAIL_REGEX.fullmatch(e):
            emails.append(e)
            seen_local.add(e)

    # Fallback: any normal emails visible on the screen
    if not emails:
        for raw in PLAIN_EMAIL_REGEX.findall(page):
            e = raw.strip().lower()
            if e and e not in seen_local:
                emails.append(e)
                seen_local.add(e)

    return emails

# =========================
# WRITE FUNCTIONS
# =========================

def write_phones_and_emails(phones, emails):
    global phones_next_row, emails_next_row, write_counter

    new_items = 0

    for p in phones:
        if p not in seen_phones:
            cell = ws_phones.cell(phones_next_row, 1)
            cell.number_format = "@"
            cell.value = p
            phones_next_row += 1
            seen_phones.add(p)
            write_counter += 1
            new_items += 1

    for e in emails:
        if e not in seen_emails:
            cell = ws_emails.cell(emails_next_row, 1)
            cell.number_format = "@"
            cell.value = e
            emails_next_row += 1
            seen_emails.add(e)
            write_counter += 1
            new_items += 1

    if new_items > 0 and write_counter % SAVE_EVERY == 0:
        maybe_save()

# =========================
# MAIN LOOP
# =========================

print("Starting in 5 seconds...")
print("Sabre must show *H9*PE*HI")
print(f"STOP KEY: {STOP_HOTKEY}")
print("IMPORTANT: Keep contacts.xlsx CLOSED while this runs.")
time.sleep(5)

try:
    while not STOP_REQUESTED:
        page = copy_sabre_screen()

        phones = extract_phones(page)
        emails = extract_emails(page)

        if phones or emails:
            write_phones_and_emails(phones, emails)

        pyautogui.typewrite("I\n")
        time.sleep(1.2)

        if STOP_REQUESTED:
            break

        pyautogui.typewrite(SABRE_COMMAND)
        time.sleep(1.5)

except KeyboardInterrupt:
    print("CTRL+C pressed, stopping...")

except SystemExit as e:
    print(str(e))

finally:
    print("Saving final data...")
    if safe_save_workbook(wb, EXCEL_FILE):
        print("Final save successful.")
    else:
        print("Final save failed. Close contacts.xlsx and retry.")

    try:
        keyboard.unhook_all_hotkeys()
    except Exception:
        pass
